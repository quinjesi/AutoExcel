import openpyxl
from openpyxl.chart import BarChart, Reference
import random

# Create a new workbook and add a worksheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sales Data for January 2025"

# Add headers to the worksheet
headers = ["Transaction ID", "Item", "Price (₦)", "Quantity", "Total Amount (₦)"]
sheet.append(headers)

# Fixed Price List
items_prices = {
    'Laptop': 300000,
    'Smartphone': 150000,
    'Tablet': 80000,
    'Headphones': 5000,
    'Smartwatch': 20000,
    'Camera': 120000,
    'Printer': 60000,
    'Monitor': 70000,
    'Keyboard': 10000,
    'Mouse': 8000,
    'Speaker': 10000,
    'Charger': 5000,
    'USB Cable': 2000,
    'External Hard Drive': 35000,
    'Webcam': 30000,
    'Power Bank': 23000,
    'Flash Drive': 4500,
    'Router': 20000,
    'Earbuds': 10000
}
items = list(items_prices.keys())

# Generate random transactions
items_total = {item: 0 for item in items}
for i in range(1, 102):
    transaction_id = f'25JAN{str(i).zfill(3)}'
    item = random.choice(items)
    price = items_prices[item]
    quantity = random.randint(1, 10)
    total_amount = price * quantity
    sheet.append([transaction_id, item, price, quantity, total_amount])

    items_total[item] += total_amount


# Total sales for the month
total_row = sheet.max_row + 2
sheet[f'D{total_row}'] = "Total Sales"
sheet[f'E{total_row}'] = f'=SUM(E2:E{sheet.max_row - 2})'


# Summary of total sales for each item
summary_start_row = total_row + 3
sheet[f'A{summary_start_row -1}'] = "Item"
sheet[f'B{summary_start_row -1}'] = "Total Sales (₦)"

row = summary_start_row
for item, total in items_total.items():
    sheet[f'A{row}'] = item
    sheet[f'B{row}'] = total
    row += 1


# Create a bar chart for total sales
chart = BarChart()
chart.title = "Total Sales for January 2025"
chart.x_axis.title = "Items"
chart.y_axis.title = "Total Amount (₦)"
data = Reference(sheet, min_col=2, min_row=summary_start_row -1, max_row=row - 1)
labels = Reference(sheet, min_col=1, min_row=summary_start_row, max_row=row - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.width = 25
chart.height = 15


# Add the chart to the worksheet
sheet.add_chart(chart, f'D{summary_start_row}')

# Save the workbook to a file
wb.save('AutoExcel.xlsx')
